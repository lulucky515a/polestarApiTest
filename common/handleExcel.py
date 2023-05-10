# -*- coding: utf-8 -*-
# @Software: PyCharm
# @File: handleExcel.py
# @Author: Luke
# @Time: 3月 31, 2023


import json
import os
import openpyxl
from common import dataDir, conFig


def createExcel(excel_name):
    """
    :param excel_name:
    :return:
    """
    wb = openpyxl.Workbook()
    wb.save(excel_name)


class DoExcel(object):

    def __init__(self, file_name=os.path.join(dataDir, conFig.getValue("Excel", "file")), sheet_name=None):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        self.wb = openpyxl.load_workbook(filename=self.file_name)
        if self.sheet_name is None:
            self.sh = self.wb.active
        else:
            self.sh = self.wb[self.sheet_name]

    def read_datas(self):
        """
        获取所有测试用例
        :param sheet_name:
        :return:
        """
        self.open()

        datas = list(self.sh.rows)
        key = [i.value for i in datas[0]]
        cases = []
        for j in datas[1:]:
            values = [c.value for c in j]
            case = dict(zip(key, values))
            cases.append(case)
        return cases

    def readData(self, row):
        """
        获取指定行数的用例
        :return:
        """
        return self.read_datas()[row-1]

    def writeData(self, row, actual, result):
        """
        写入数据
        :param row:
        :return:
        """
        self.open()
        if isinstance(row, int) and (2 <= row <= self.sh.max_row):
            self.sh.cell(row, 7).value = actual
            self.sh.cell(row, 8).value = result
            self.wb.save(filename=self.file_name)
        else:
            print('传入的行号有误，行号应为大于1的整数！')

    @staticmethod
    def assertDictItem(dic1, dic2):
        """
        断言dic1中的所有元素都是dict2中的成员，成立返回True,不成立引发断言错误
        :param dic1: 字典
        :param dic2: 字典
        :return:
        """
        for item in dic1.items():
            if item not in dic2.items():
                raise AssertionError("{} items not in {}".format(dic1, dic2))

    @staticmethod
    def check_json_keys(actual_json, expected_json):
        try:
            actual_keys = set(json.loads(actual_json).keys())
            expected_keys = set(json.loads(expected_json).keys())
        except (json.decoder.JSONDecodeError, ValueError) as e:
            print(f"Invalid JSON data: {e}")
            return False
        if actual_keys != expected_keys:
            print(f"Actual keys {actual_keys} do not match expected keys {expected_keys}")
            return False
        return True

    def writeDatas(self, row, orderNumber, orderID, carType, financeOrder, price,
                   finance20006, finance20009, finance20013):
        self.open()
        if isinstance(row, int) and (2 <= row <= self.sh.max_row):
            self.sh.cell(row, 3).value = orderNumber
            self.sh.cell(row, 4).value = orderID
            self.sh.cell(row, 5).value = carType
            self.sh.cell(row, 6).value = financeOrder
            self.sh.cell(row, 7).value = price
            self.sh.cell(row, 8).value = finance20006
            self.sh.cell(row, 9).value = finance20009
            self.sh.cell(row, 10).value = finance20013
            self.wb.save(filename=self.file_name)
        else:
            print('传入的行号有误，行号应为大于1的整数！')


if __name__ == '__main__':
    valorant = DoExcel(file_name=r"/Users/luzhixiang/PycharmProjects/polestarApiTest/datas/bpmPr.xlsx",
                       sheet_name="profile")
    list1 = valorant.read_datas()
    # print(list1)
    data = """
    {
    "id": "ebfdefd9-092b-4575-9530-3a4947333661",
    "avatar": "https://pscnid-profile-avatar-stg.s3.cn-northwest-1.amazonaws.com.cn/avatar-ebfdefd9-092b-4575-9530-3a4947333661-1683341862475.jpeg",
    "nickname": "lu",
    "mobile": "13816031063",
    "birthday": null,
    "birthdaymodifycount": 0,
    "gender": "",
    "focusmodel": "",
    "hobbies": "{\"科技\",\"运动\",\"时尚\"}",
    "province": "",
    "city": "",
    "district": "",
    "address": "",
    "detail": null
}
    """
    print(list1[0]["expected"])
    print(valorant.check_json_keys(data, json.dumps(list1[0]["expected"])))